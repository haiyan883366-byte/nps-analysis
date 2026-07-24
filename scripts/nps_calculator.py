#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
NPS分析与转化报告生成脚本 v4 (统一版)
========================================================
  - 双底表模式（推荐）: --nps-file + --xingke-file → 自动合并 → NPS分析
  - 单底表模式（兼容）: --nps-file → 直接 NPS分析
  - 辅导维度 + 组长维度双表输出
  - Excel 带 ColorScale 色阶（3 sheet）
  - HTML 报告带颜色标注
  - NPS口径: 推荐者>=10, 贬损者<=7 (底表1-11分制)
  - 数据校验: nps列值超出1-11时告警
"""

import pandas as pd
import sys
import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════
# 列名适配
# ═══════════════════════════════════════════════════════════

def _find_col(df, candidates, default=None):
    """在DataFrame列中查找第一个匹配的列名"""
    for c in candidates:
        if c in df.columns:
            return c
    if default is not None:
        return default
    raise KeyError(f"未找到所需列，候选: {candidates}")


def _normalize_columns(df):
    """统一列名，支持多种底表格式"""
    mapping = {}
    rec_col = _find_col(df, ['【推荐值】', '推荐值', 'nps'])
    if rec_col != '推荐值':
        mapping[rec_col] = '推荐值'
    reg_col = _find_col(df, ['是否考虑报名长期班', '是否报名'])
    if reg_col != '是否报名':
        mapping[reg_col] = '是否报名'
    grade_col = _find_col(df, ['年级', 'grade_names'])
    if grade_col != '年级':
        mapping[grade_col] = '年级'
    tutor_col = _find_col(df, ['辅导', 'counselor_name', '辅导姓名'])
    if tutor_col != '辅导':
        mapping[tutor_col] = '辅导'
    leader_col = _find_col(df, ['组长', '组长姓名'])
    if leader_col != '组长':
        mapping[leader_col] = '组长'
    if mapping:
        df = df.rename(columns=mapping)
    return df


# ═══════════════════════════════════════════════════════════
# 数据校验
# ═══════════════════════════════════════════════════════════

def _validate_nps_column(df, col='推荐值'):
    """
    校验 NPS 推荐值列是否在预期范围 1-11 内。
    如果存在超出范围的值，打印告警信息。
    返回: (是否正常, 异常数量, 异常值列表)
    """
    if col not in df.columns:
        return True, 0, []

    vals = df[col].dropna()
    invalid_mask = (vals < 1) | (vals > 11)
    invalid_count = invalid_mask.sum()

    if invalid_count > 0:
        invalid_vals = vals[invalid_mask].value_counts().to_dict()
        print(f'\n{"="*60}')
        print(f'⚠️  ⚠️  ⚠️  数据质量告警  ⚠️  ⚠️  ⚠️')
        print(f'{"="*60}')
        print(f'  NPS推荐值列 "{col}" 存在 {invalid_count} 条超出 1-11 范围的异常值')
        print(f'  异常值分布: {invalid_vals}')
        print(f'  总样本量: {len(df)}, 异常占比: {invalid_count/len(df)*100:.1f}%')
        print(f'  ⚠️  请检查底表数据是否正确！')
        print(f'  异常行将被自动剔除，NPS计算仅使用1-11分数据。')
        print(f'{"="*60}\n')
        return False, invalid_count, invalid_vals

    return True, 0, []


# ═══════════════════════════════════════════════════════════
# 两表合并（双底表模式）
# ═══════════════════════════════════════════════════════════

# NPS转化底表列名重命名映射
NPS_RENAME_MAP = {
    'q1': 'nps',
    'q2': 'q2喜爱度',
    'q3': 'q3收获感',
    'q4': 'q4专注力',
    'q5': 'q5表达意愿',
    'q10': '是否考虑报名长期班',
}


def _auto_detect_sheet(file_path, required_cols, default_sheet=0):
    """
    自动检测包含所需列的 sheet。
    如果 default_sheet 已包含所需列，直接使用；
    否则遍历所有 sheet 找到包含最多所需列的那个。
    """
    xls = pd.ExcelFile(file_path)
    df_try = pd.read_excel(file_path, sheet_name=default_sheet, nrows=2)
    found = sum(1 for c in required_cols if c in df_try.columns)
    if found >= len(required_cols) * 0.5:
        return default_sheet

    best_sheet = default_sheet
    best_count = 0
    for s in xls.sheet_names:
        df_s = pd.read_excel(file_path, sheet_name=s, nrows=2)
        cnt = sum(1 for c in required_cols if c in df_s.columns)
        if cnt > best_count:
            best_count = cnt
            best_sheet = s
    return best_sheet


def merge_tables(nps_file, xingke_file, nps_sheet=0, xk_sheet=0):
    """
    合并 NPS转化底表 + 行课底表

    步骤:
      1. 读取两个底表（自动检测 sheet）
      2. 重命名 NPS转化底表列名 (q1→nps 等)
      3. 三重键匹配教学组 (辅导名+年级+课程ID)
      4. 降级匹配 (辅导名+年级)
      5. 填充组长姓名列

    返回: 合并后的完整 DataFrame（未筛选年级）
    """
    # ── 1. 读取两个底表（自动检测 sheet）──
    print("  读取底表...")

    nps_required = ['counselor_name', 'grade_names', 'fk_course', '辅导', '年级']
    actual_nps_sheet = _auto_detect_sheet(nps_file, nps_required, nps_sheet)
    if actual_nps_sheet != nps_sheet:
        print(f"    NPS底表: 自动切换到 sheet '{actual_nps_sheet}'")
    df_nps = pd.read_excel(nps_file, sheet_name=actual_nps_sheet)

    xk_required = ['辅导名', '年级', '课程ID', '教学组', 'counselor_name']
    actual_xk_sheet = _auto_detect_sheet(xingke_file, xk_required, xk_sheet)
    if actual_xk_sheet != xk_sheet:
        print(f"    行课底表: 自动切换到 sheet '{actual_xk_sheet}'")
    df_xk = pd.read_excel(xingke_file, sheet_name=actual_xk_sheet)

    print(f"    NPS转化底表: {len(df_nps)} 行  行课底表: {len(df_xk)} 行")

    # ── 2. 重命名NPS转化底表列名 ──
    rename_map = {}
    for old, new in NPS_RENAME_MAP.items():
        if old in df_nps.columns:
            rename_map[old] = new
    if rename_map:
        df_nps = df_nps.rename(columns=rename_map)
        print(f"    已重命名: {rename_map}")

    # ── 3. 识别列名 ──
    nps_tutor_col = _find_col(df_nps, ['counselor_name', '辅导', '辅导姓名'], 'counselor_name')
    nps_grade_col = _find_col(df_nps, ['grade_names', '年级'], 'grade_names')
    nps_course_col = _find_col(df_nps, ['fk_course', '课程ID', 'course_id'], 'fk_course')

    xk_tutor_col = _find_col(df_xk, ['辅导名', '辅导', 'counselor_name', '辅导姓名'], '辅导名')
    xk_grade_col = _find_col(df_xk, ['年级', 'grade_names'], '年级')
    xk_course_col = _find_col(df_xk, ['课程ID', 'fk_course', 'course_id'], '课程ID')
    xk_group_col = _find_col(df_xk, ['教学组', '组长', '组长姓名'], '教学组')

    # ── 4. 提取教学组映射 ──
    xk_3key = df_xk.groupby([xk_tutor_col, xk_grade_col, xk_course_col])[xk_group_col].first().reset_index()
    xk_2key = df_xk.groupby([xk_tutor_col, xk_grade_col])[xk_group_col].first().reset_index()

    # ── 5. 三重键匹配 ──
    df_nps['_mk3'] = list(zip(df_nps[nps_tutor_col], df_nps[nps_grade_col], df_nps[nps_course_col]))
    xk_3key['_mk3'] = list(zip(xk_3key[xk_tutor_col], xk_3key[xk_grade_col], xk_3key[xk_course_col]))
    dict_3key = dict(zip(xk_3key['_mk3'], xk_3key[xk_group_col]))
    df_nps['教学组'] = df_nps['_mk3'].map(dict_3key)

    matched_3 = df_nps['教学组'].notna().sum()
    pct_3 = matched_3 / len(df_nps) * 100 if len(df_nps) > 0 else 0
    print(f"    三重键匹配: {matched_3}/{len(df_nps)} ({pct_3:.1f}%)")

    # ── 6. 降级匹配 ──
    unmatched = df_nps['教学组'].isna()
    if unmatched.any():
        dict_2key = dict(zip(
            zip(xk_2key[xk_tutor_col], xk_2key[xk_grade_col]),
            xk_2key[xk_group_col]
        ))
        fb_keys = list(zip(df_nps.loc[unmatched, nps_tutor_col], df_nps.loc[unmatched, nps_grade_col]))
        fb_matched = pd.Series(fb_keys).map(dict_2key)
        df_nps.loc[unmatched, '教学组'] = fb_matched.values
        matched_2 = df_nps['教学组'].notna().sum() - matched_3
        print(f"    降级匹配: +{matched_2}")

    still_unmatched = df_nps['教学组'].isna().sum()
    pct_un = still_unmatched / len(df_nps) * 100 if len(df_nps) > 0 else 0
    print(f"    最终未匹配: {still_unmatched}/{len(df_nps)} ({pct_un:.1f}%)")

    if still_unmatched > 0:
        unmatched_tutors = df_nps[df_nps['教学组'].isna()][nps_tutor_col].unique()
        print(f"    未匹配辅导: {list(unmatched_tutors)[:15]}")

    # ── 7. 填充组长姓名列 ──
    if '组长姓名' not in df_nps.columns:
        df_nps['组长姓名'] = None
    df_nps['组长姓名'] = df_nps['教学组']

    # 清理临时列
    df_nps = df_nps.drop(columns=['_mk3'], errors='ignore')

    # 关键统计
    print(f"    合并结果: {len(df_nps)} 行, "
          f"{df_nps[nps_tutor_col].nunique()} 位辅导, "
          f"{df_nps['教学组'].nunique()} 个教学组")

    return df_nps


# ═══════════════════════════════════════════════════════════
# NPS 计算
# ═══════════════════════════════════════════════════════════

def calc_nps(group):
    """对一组用户计算NPS和转化指标

    NPS口径（底表1-11分制）:
      - 推荐者: >=10 (对应标准0-10量表的9-10)
      - 中立者: 8-9 (对应标准0-10量表的7-8)
      - 贬损者: <=7 (对应标准0-10量表的0-6)
      - NPS = 推荐者占比 - 贬损者占比
    """
    total = len(group)
    promoters = len(group[group['推荐值'] >= 10])
    detractors = len(group[group['推荐值'] <= 7])
    nps = (promoters - detractors) / total * 100
    return pd.Series({
        '样本量': total,
        'NPS': round(nps, 1),
        '推荐率%': round(promoters / total * 100, 1),
        '贬损率%': round(detractors / total * 100, 1),
        '报名率%': round(len(group[group['是否报名'] == 1]) / total * 100, 1),
        '考虑率%': round(len(group[group['是否报名'] == 2]) / total * 100, 1),
        '拒绝率%': round(len(group[group['是否报名'] == 3]) / total * 100, 1),
    })


# ═══════════════════════════════════════════════════════════
# Excel 生成（带色阶 + 多 sheet）
# ═══════════════════════════════════════════════════════════

HDR_FONT   = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
HDR_FILL   = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HDR_ALIGN  = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_FONT  = Font(name='Microsoft YaHei', size=11)
CELL_ALIGN = Alignment(horizontal='center', vertical='center')
CELL_LEFT  = Alignment(horizontal='left', vertical='center')
BORDER     = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

COLOR_SCALES = {
    'NPS':      ('正向', '00B050', 'FFFFFF', 'FF6666'),
    '推荐率%':  ('正向', '2F5496', 'FFFFFF', 'D6E4F0'),
    '贬损率%':  ('反向', 'FF4444', 'FFFFFF', '00B050'),
    '报名率%':  ('正向', '00B050', 'FFFFFF', 'FFC000'),
    '拒绝率%':  ('反向', 'FF4444', 'FFFFFF', '548235'),
    '考虑率%':  ('中性', None, None, None),
}

TUTOR_HEADERS  = ['辅导老师', '组长', '样本量', 'NPS', '推荐率%', '贬损率%', '报名率%', '考虑率%', '拒绝率%']
TUTOR_WIDTHS   = [14, 12, 10, 10, 10, 10, 10, 10, 10]
LEADER_HEADERS = ['组长', '样本量', 'NPS', '推荐率%', '贬损率%', '报名率%', '考虑率%', '拒绝率%']
LEADER_WIDTHS  = [14, 10, 10, 10, 10, 10, 10, 10]


def _write_sheet(ws, headers, data_rows, col_widths):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = BORDER
    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT
            cell.alignment = CELL_LEFT if c == 1 else CELL_ALIGN
            cell.border = BORDER
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'


def _add_color_scales(ws, data_start_row, data_end_row, headers):
    n = data_end_row - data_start_row + 1
    if n <= 1:
        return
    for col_name in headers:
        cfg = COLOR_SCALES.get(col_name)
        if cfg is None or cfg[0] == '中性':
            continue
        direction, hi_color, mid_color, lo_color = cfg
        try:
            ci = headers.index(col_name) + 1
        except ValueError:
            continue
        col_letter = get_column_letter(ci)
        rng = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
        if direction == '正向':
            ws.conditional_formatting.add(rng,
                ColorScaleRule(start_type='min', start_color=lo_color,
                               mid_type='percentile', mid_value=50, mid_color=mid_color,
                               end_type='max', end_color=hi_color))
        else:
            ws.conditional_formatting.add(rng,
                ColorScaleRule(start_type='min', start_color=hi_color,
                               mid_type='percentile', mid_value=50, mid_color=mid_color,
                               end_type='max', end_color=lo_color))


def _build_tutor_rows(df_result):
    rows = []
    for _, row in df_result.iterrows():
        rows.append([
            row['辅导'], row['组长'], int(row['样本量']),
            row['NPS'], row['推荐率%'], row['贬损率%'],
            row['报名率%'], row['考虑率%'], row['拒绝率%'],
        ])
    return rows


def _build_leader_rows(df_leader):
    rows = []
    for _, row in df_leader.iterrows():
        rows.append([
            row['组长'], int(row['样本量']),
            row['NPS'], row['推荐率%'], row['贬损率%'],
            row['报名率%'], row['考虑率%'], row['拒绝率%'],
        ])
    return rows


def _generate_excel_with_colors(result_tutor, result_leader, grade_label, output_dir):
    wb = Workbook()
    tutor_rows = _build_tutor_rows(result_tutor)
    leader_rows = _build_leader_rows(result_leader)

    ws1 = wb.active
    ws1.title = "辅导维度"
    _write_sheet(ws1, TUTOR_HEADERS, tutor_rows, TUTOR_WIDTHS)
    _add_color_scales(ws1, 2, len(tutor_rows) + 1, TUTOR_HEADERS)

    ws2 = wb.create_sheet("组长维度")
    _write_sheet(ws2, LEADER_HEADERS, leader_rows, LEADER_WIDTHS)
    _add_color_scales(ws2, 2, len(leader_rows) + 1, LEADER_HEADERS)

    # Sheet 3: 汇总对比
    ws3 = wb.create_sheet("汇总对比")
    title1_row = 1
    ws3.merge_cells(start_row=title1_row, start_column=1, end_row=title1_row, end_column=len(TUTOR_HEADERS))
    ws3.cell(row=title1_row, column=1, value=f'{grade_label} — 辅导维度 NPS（按NPS降序）').font = \
        Font(name='Microsoft YaHei', bold=True, size=13, color='2F5496')

    tutor_hdr_row = title1_row + 1
    for c, h in enumerate(TUTOR_HEADERS, 1):
        cell = ws3.cell(row=tutor_hdr_row, column=c, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = HDR_ALIGN; cell.border = BORDER
    for r, row_data in enumerate(tutor_rows):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r + tutor_hdr_row + 1, column=c, value=val)
            cell.font = CELL_FONT
            cell.alignment = CELL_LEFT if c == 1 else CELL_ALIGN
            cell.border = BORDER
    tutor_end_row = tutor_hdr_row + len(tutor_rows)
    _add_color_scales(ws3, tutor_hdr_row + 1, tutor_end_row, TUTOR_HEADERS)

    leader_title_row = tutor_end_row + 3
    ws3.merge_cells(start_row=leader_title_row, start_column=1, end_row=leader_title_row, end_column=len(LEADER_HEADERS))
    ws3.cell(row=leader_title_row, column=1, value=f'{grade_label} — 组长维度汇总').font = \
        Font(name='Microsoft YaHei', bold=True, size=13, color='2F5496')

    leader_hdr_row = leader_title_row + 1
    for c, h in enumerate(LEADER_HEADERS, 1):
        cell = ws3.cell(row=leader_hdr_row, column=c, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = HDR_ALIGN; cell.border = BORDER
    for r, row_data in enumerate(leader_rows):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r + leader_hdr_row + 1, column=c, value=val)
            cell.font = CELL_FONT
            cell.alignment = CELL_LEFT if c == 1 else CELL_ALIGN
            cell.border = BORDER
    leader_end_row = leader_hdr_row + len(leader_rows)
    _add_color_scales(ws3, leader_hdr_row + 1, leader_end_row, LEADER_HEADERS)

    for c, w in enumerate(TUTOR_WIDTHS, 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    excel_path = os.path.join(output_dir, f'{grade_label}NPS分析表.xlsx')
    wb.save(excel_path)
    print(f'  Excel: {excel_path} ({len(tutor_rows)} 位辅导, {len(leader_rows)} 位组长)')
    return excel_path


# ═══════════════════════════════════════════════════════════
# HTML 生成
# ═══════════════════════════════════════════════════════════

def _html_section_table(rows_html, headers, section_title, note_text=''):
    th_html = ''.join(f'<th>{h}</th>' for h in headers)
    return f'''
    <div class="section">
        <h2>{section_title}</h2>
        {note_text}
        <table><thead><tr>{th_html}</tr></thead><tbody>
        {''.join(rows_html)}
        </tbody></table>
    </div>'''


def _build_tutor_html_rows(result_tutor):
    rows = []
    for _, row in result_tutor.iterrows():
        nps = row['NPS']
        rec = row['推荐率%']
        det = row['贬损率%']
        sig = row['报名率%']
        nos = row['拒绝率%']

        if nps >= 60: nbg, nc = '#2e7d32', 'white'
        elif nps >= 50: nbg, nc = '#43a047', 'white'
        elif nps >= 40: nbg, nc = '#66bb6a', 'white'
        elif nps >= 30: nbg, nc = '#a5d6a7', '#1b5e20'
        else: nbg, nc = '#c8e6c9', '#1b5e20'

        if rec >= 75: rbg, rc = '#1976d2', 'white'
        elif rec >= 70: rbg, rc = '#1e88e5', 'white'
        elif rec >= 65: rbg, rc = '#42a5f5', 'white'
        elif rec >= 60: rbg, rc = '#64b5f6', '#0d47a1'
        else: rbg, rc = '#90caf9', '#0d47a1'

        if det >= 22: dbg, dc = '#1976d2', 'white'
        elif det >= 18: dbg, dc = '#1e88e5', 'white'
        elif det >= 15: dbg, dc = '#42a5f5', 'white'
        elif det >= 12: dbg, dc = '#64b5f6', '#0d47a1'
        else: dbg, dc = '#90caf9', '#0d47a1'

        if sig >= 18: sbg, sc = '#2e7d32', 'white'
        elif sig >= 15: sbg, sc = '#43a047', 'white'
        elif sig >= 12: sbg, sc = '#66bb6a', '#1b5e20'
        elif sig >= 10: sbg, sc = '#a5d6a7', '#1b5e20'
        elif sig >= 8: sbg, sc = '#ffcc80', '#e65100'
        elif sig >= 5: sbg, sc = '#ffab91', '#bf360c'
        else: sbg, sc = '#ef9a9a', '#b71c1c'

        if nos >= 15: nbg2, nc2 = '#1976d2', 'white'
        elif nos >= 12: nbg2, nc2 = '#1e88e5', 'white'
        elif nos >= 8: nbg2, nc2 = '#42a5f5', 'white'
        elif nos >= 5: nbg2, nc2 = '#64b5f6', '#0d47a1'
        else: nbg2, nc2 = '#90caf9', '#0d47a1'

        rows.append(f'''<tr>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:left;font-weight:500;">{row['辅导']}</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:left;">{row['组长']}</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;background:{nbg};color:{nc};font-weight:bold;">{nps:.1f}%</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;background:{rbg};color:{rc};">{rec:.1f}%</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;background:{dbg};color:{dc};">{det:.1f}%</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;background:{sbg};color:{sc};font-weight:bold;">{sig:.1f}%</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;">{row['考虑率%']:.1f}%</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;background:{nbg2};color:{nc2};">{nos:.1f}%</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;color:#666;">{row['样本量']:.0f}</td>
        </tr>''')
    return rows


def _build_leader_html_rows(result_leader):
    rows = []
    for _, row in result_leader.iterrows():
        nps = row['NPS']
        rec = row['推荐率%']
        det = row['贬损率%']
        sig = row['报名率%']
        nos = row['拒绝率%']

        if nps >= 60: nbg, nc = '#2e7d32', 'white'
        elif nps >= 50: nbg, nc = '#43a047', 'white'
        elif nps >= 40: nbg, nc = '#66bb6a', 'white'
        elif nps >= 30: nbg, nc = '#a5d6a7', '#1b5e20'
        else: nbg, nc = '#c8e6c9', '#1b5e20'

        if rec >= 75: rbg, rc = '#1976d2', 'white'
        elif rec >= 70: rbg, rc = '#1e88e5', 'white'
        elif rec >= 65: rbg, rc = '#42a5f5', 'white'
        elif rec >= 60: rbg, rc = '#64b5f6', '#0d47a1'
        else: rbg, rc = '#90caf9', '#0d47a1'

        if det >= 22: dbg, dc = '#1976d2', 'white'
        elif det >= 18: dbg, dc = '#1e88e5', 'white'
        elif det >= 15: dbg, dc = '#42a5f5', 'white'
        elif det >= 12: dbg, dc = '#64b5f6', '#0d47a1'
        else: dbg, dc = '#90caf9', '#0d47a1'

        if sig >= 18: sbg, sc = '#2e7d32', 'white'
        elif sig >= 15: sbg, sc = '#43a047', 'white'
        elif sig >= 12: sbg, sc = '#66bb6a', '#1b5e20'
        elif sig >= 10: sbg, sc = '#a5d6a7', '#1b5e20'
        elif sig >= 8: sbg, sc = '#ffcc80', '#e65100'
        elif sig >= 5: sbg, sc = '#ffab91', '#bf360c'
        else: sbg, sc = '#ef9a9a', '#b71c1c'

        if nos >= 15: nbg2, nc2 = '#1976d2', 'white'
        elif nos >= 12: nbg2, nc2 = '#1e88e5', 'white'
        elif nos >= 8: nbg2, nc2 = '#42a5f5', 'white'
        elif nos >= 5: nbg2, nc2 = '#64b5f6', '#0d47a1'
        else: nbg2, nc2 = '#90caf9', '#0d47a1'

        rows.append(f'''<tr>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:left;font-weight:500;">{row['组长']}</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;color:#666;">{row['样本量']:.0f}</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;background:{nbg};color:{nc};font-weight:bold;">{nps:.1f}%</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;background:{rbg};color:{rc};">{rec:.1f}%</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;background:{dbg};color:{dc};">{det:.1f}%</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;background:{sbg};color:{sc};font-weight:bold;">{sig:.1f}%</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;">{row['考虑率%']:.1f}%</td>
            <td style="padding:10px 14px;border-bottom:1px solid #e8eaed;text-align:center;background:{nbg2};color:{nc2};">{nos:.1f}%</td>
        </tr>''')
    return rows


def generate_html(result_tutor, result_leader, total_samples, grade_name, date_label):
    tutor_rows = _build_tutor_html_rows(result_tutor)
    leader_rows = _build_leader_html_rows(result_leader)
    tutor_section = _html_section_table(
        tutor_rows, TUTOR_HEADERS,
        f'{grade_name} — 辅导维度 NPS（按 NPS 降序）',
        f'<p class="note">样本量：{total_samples} 份 | {len(result_tutor)} 位辅导老师</p>')
    leader_section = _html_section_table(leader_rows, LEADER_HEADERS, f'{grade_name} — 组长维度汇总')

    return f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
    body {{ font-family: "Microsoft YaHei", "PingFang SC", Arial, sans-serif; background: #f8f9fa; padding: 30px; margin: 0; }}
    .container {{ max-width: 1050px; margin: 0 auto; background: white; padding: 32px; border-radius: 12px; box-shadow: 0 4px 20px rgba(0,0,0,0.08); }}
    h1 {{ font-size: 22px; color: #1a1a1a; margin-bottom: 8px; text-align: center; font-weight: 600; }}
    .subtitle {{ font-size: 13px; color: #888; margin-bottom: 30px; text-align: center; }}
    .section {{ margin-bottom: 36px; }}
    .section h2 {{ font-size: 16px; color: #2F5496; margin-bottom: 6px; border-left: 4px solid #4472C4; padding-left: 10px; }}
    .note {{ font-size: 12px; color: #999; margin-bottom: 8px; }}
    table {{ width: 100%; border-collapse: separate; border-spacing: 0; font-size: 13px; }}
    th {{ padding: 12px 14px; background: #f0f4f8; color: #333; font-weight: 600; text-align: center; border-bottom: 2px solid #d0d7de; border-top: 1px solid #e8eaed; font-size: 13px; }}
    th:first-child {{ border-top-left-radius: 8px; text-align: left; }}
    th:last-child {{ border-top-right-radius: 8px; }}
    tr:last-child td:first-child {{ border-bottom-left-radius: 8px; }}
    tr:last-child td:last-child {{ border-bottom-right-radius: 8px; }}
    tr:hover td {{ filter: brightness(0.97); }}
    td {{ transition: all 0.15s ease; }}
</style></head><body>
<div class="container">
<h1>{grade_name} NPS 与转化分析报告</h1>
<div class="subtitle">数据时间：{date_label}</div>
{tutor_section}
{leader_section}
</div></body></html>'''


# ═══════════════════════════════════════════════════════════
# 主入口（v4 统一版：支持单底表/双底表）
# ═══════════════════════════════════════════════════════════

def run(nps_file, xingke_file=None, grade=None, min_samples=10, output_dir=None,
        nps_sheet=0, xk_sheet=0):
    """
    NPS 分析主函数 v4 —— 双底表 / 单底表 统一入口。

    双底表模式（推荐）:
      传入 nps_file + xingke_file → 自动合并 → 年级筛选 → NPS分析 → Excel+HTML

    单底表模式（兼容）:
      只传 nps_file → 直接列名适配 → NPS分析 → Excel+HTML

    参数:
        nps_file:    NPS转化底表路径（必填）
        xingke_file: 行课底表路径（可选）。提供时走双底表合并流程。
        grade:       筛选年级，None 表示全量
        min_samples: 最小样本量阈值，默认 10
        output_dir:  输出目录，None 则输出到 nps_file 所在目录
        nps_sheet:   NPS底表 sheet 名或索引，默认 0
        xk_sheet:    行课底表 sheet 名或索引，默认 0

    返回: (result_tutor, result_leader, [merged_path])
    """
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(nps_file)) or '.'

    # ── Phase 0: 双底表合并（可选）──
    merged_path = None
    if xingke_file:
        print("=" * 60)
        print("Phase 0: 两表合并")
        print("=" * 60)
        df_merged = merge_tables(nps_file, xingke_file, nps_sheet, xk_sheet)

        # 年级筛选
        grade_col = _find_col(df_merged, ['grade_names', '年级'], 'grade_names')
        if grade and grade in df_merged[grade_col].unique():
            df_input = df_merged[df_merged[grade_col] == grade].copy()
            grade_label = grade
        else:
            df_input = df_merged.copy()
            grade_label = '全年级'
            if grade:
                print(f"  ⚠️ 未找到年级「{grade}」，使用全量数据")

        print(f"\n  {grade_label}: {len(df_input)} 行")

        # 保存合并底表
        merged_path = os.path.join(output_dir, f'{grade_label}合并后底表.xlsx')
        df_input.to_excel(merged_path, index=False, sheet_name=f'{grade_label}NPS底表')
        print(f"  合并底表: {merged_path}\n")

        # 后续分析使用合并后的底表
        analysis_file = merged_path
        analysis_sheet = 0
        analysis_grade = None  # 已经在合并时筛选过了

    else:
        # 单底表模式
        print("=" * 60)
        print("单底表模式: 直接分析")
        print("=" * 60)
        analysis_file = nps_file
        analysis_sheet = nps_sheet
        analysis_grade = grade
        if grade:
            grade_label = grade
        else:
            grade_label = '全年级'

    # ── 分析 ──
    print("=" * 60)
    print(f"Phase {'1' if xingke_file else ''}: NPS 分析 ({grade_label})")
    print("=" * 60)

    df = pd.read_excel(analysis_file, sheet_name=analysis_sheet)
    df = _normalize_columns(df)

    # 数据校验
    is_valid, inv_count, inv_vals = _validate_nps_column(df, '推荐值')
    if not is_valid:
        valid_mask = (df['推荐值'] >= 1) & (df['推荐值'] <= 11) | df['推荐值'].isna()
        df = df[valid_mask].copy()

    # 如果单底表且指定了年级，筛选
    if analysis_grade and analysis_grade in df['年级'].unique():
        df = df[df['年级'] == analysis_grade].copy()

    total = len(df)

    # 辅导维度
    result_tutor = (
        df.groupby(['辅导', '组长'], dropna=False)
        .apply(calc_nps, include_groups=False)
        .reset_index()
    )
    result_tutor = result_tutor[result_tutor['样本量'] >= min_samples].copy()
    result_tutor['组长'] = result_tutor['组长'].fillna('未分组')
    result_tutor = result_tutor.sort_values(['组长', 'NPS'], ascending=[True, False]).reset_index(drop=True)

    # 组长维度
    result_leader = (
        df.groupby('组长')
        .apply(calc_nps, include_groups=False)
        .reset_index()
    )
    result_leader['组长'] = result_leader['组长'].fillna('未分组')
    result_leader = result_leader.sort_values('NPS', ascending=False).reset_index(drop=True)

    # 输出 Excel
    _generate_excel_with_colors(result_tutor, result_leader, grade_label, output_dir)

    # 输出 HTML
    date_label = pd.Timestamp.now().strftime('%Y年%m月')
    html = generate_html(result_tutor, result_leader, total, grade_label, date_label)
    html_path = os.path.join(output_dir, f'{grade_label}NPS分析表.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'  HTML: {html_path}')

    # 汇总
    print(f"\n===== 完成 =====")
    print(f"  {grade_label}: {total} 份样本, {len(result_tutor)} 位辅导, {len(result_leader)} 位组长")
    if merged_path:
        print(f"  1. 合并底表: {merged_path}")
    print(f"  {'2' if merged_path else '1'}. NPS汇总表: {os.path.join(output_dir, f'{grade_label}NPS分析表.xlsx')}")
    print(f"  {'3' if merged_path else '2'}. HTML报告:  {html_path}")

    return result_tutor, result_leader


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='NPS分析与转化报告生成 v4（双底表/单底表统一）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 双底表模式（推荐）
  python nps_calculator.py --nps-file "NPS转化底表.xlsx" --xingke-file "行课底表.xlsx" --grade "一年级"
  # 单底表模式（兼容）
  python nps_calculator.py "NPS问卷表.xlsx" --grade "一年级"
        """)
    parser.add_argument('input_file', nargs='?', default=None, help='Excel文件路径（单底表模式）')
    parser.add_argument('--nps-file', default=None, help='NPS转化底表路径')
    parser.add_argument('--xingke-file', default=None, help='行课底表路径（可选，提供时走双底表合并流程）')
    parser.add_argument('--grade', '-g', default=None, help='筛选年级')
    parser.add_argument('--min-samples', '-m', type=int, default=10, help='最小样本量阈值（默认10）')
    parser.add_argument('--output-dir', '-o', default=None, help='输出目录')
    parser.add_argument('--nps-sheet', default=0, help='NPS底表sheet名或索引（默认0）')
    parser.add_argument('--xk-sheet', default=0, help='行课底表sheet名或索引（默认0）')
    args = parser.parse_args()

    # 确定 nps_file
    nps_file = args.nps_file or args.input_file
    if not nps_file:
        parser.error("必须提供 --nps-file 或位置参数 input_file")

    run(
        nps_file=nps_file,
        xingke_file=args.xingke_file,
        grade=args.grade,
        min_samples=args.min_samples,
        output_dir=args.output_dir,
        nps_sheet=args.nps_sheet,
        xk_sheet=args.xk_sheet,
    )
